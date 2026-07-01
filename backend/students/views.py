from rest_framework import generics, filters
from rest_framework.exceptions import PermissionDenied
from config.pagination import StandardResultsSetPagination
from .models import Student
from .serializers import StudentSerializer
from accounts.permissions import IsAdmin, IsAnyAuthenticated, IsAdminOrFaculty


class StudentListCreateView(generics.ListCreateAPIView):
    serializer_class = StudentSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'email', 'register_number', 'course']

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAnyAuthenticated()]
        return [IsAdminOrFaculty()]

    def get_queryset(self):
        user = self.request.user
        queryset = Student.objects.all()
        
        if user.role == 'student':
            queryset = Student.objects.filter(id=user.linked_student_id) if user.linked_student_id else Student.objects.none()
        elif user.role == 'faculty':
            queryset = queryset.filter(department=user.department)
            
        course = self.request.query_params.get('course')
        if course:
            queryset = queryset.filter(course=course)
            
        department = self.request.query_params.get('department')
        if department:
            queryset = queryset.filter(department=department)
            
        year = self.request.query_params.get('year')
        if year:
            queryset = queryset.filter(year=year)
            
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'faculty':
            student = serializer.save(department=user.department)
        else:
            student = serializer.save()
            
        # Auto-link to existing User account if it exists
        from accounts.models import User
        target_user = User.objects.filter(email=student.email, role='student').first()
        if target_user and not target_user.linked_student:
            target_user.linked_student = student
            target_user.save()


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = StudentSerializer

    def get_permissions(self):
        if self.request.method in ['GET', 'PUT', 'PATCH']:
            return [IsAnyAuthenticated()]
        return [IsAdminOrFaculty()]

    def perform_update(self, serializer):
        user = self.request.user
        if user.role in ['student', 'parent']:
            unsafe_fields = ['first_name', 'last_name', 'email', 'register_number', 'department', 'course', 'year', 'cgpa', 'attendance_percentage', 'status']
            for f in unsafe_fields:
                serializer.validated_data.pop(f, None)
        serializer.save()

    def get_queryset(self):
        return Student.objects.all()

    def get_object(self):
        obj = super().get_object()
        user = self.request.user
        if user.role == 'student' and user.linked_student_id != obj.id:
            raise PermissionDenied('You can only view your own profile.')
        if user.role == 'faculty' and obj.department != user.department:
            raise PermissionDenied('You can only manage students in your department.')
        return obj

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny
import openpyxl

class StudentBulkUploadSampleView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Register Number', 'Department', 'Course', 'Year', 'Gender']
        ws.append(headers)

        example = ['John', 'Doe', 'john.doe@example.com', '1234567890', 'REG001', 'Computer Science', 'B.Tech - Computer Science & Engineering', 1, 'male']
        ws.append(example)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=student_bulk_upload_sample.xlsx'
        wb.save(response)
        return response

class StudentBulkUploadView(APIView):
    permission_classes = [IsAdminOrFaculty]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file uploaded'}, status=400)
            
        if not file_obj.name.endswith('.xlsx'):
            return Response({'error': 'Only .xlsx files are supported'}, status=400)

        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Failed to parse excel file: {str(e)}'}, status=400)

        errors = []
        created_count = 0
        
        user = request.user
        enforced_dept = user.department if user.role == 'faculty' else None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx <= 2:
                continue
                
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
                
            try:
                first_name = str(row[0]).strip() if row[0] else ''
                last_name = str(row[1]).strip() if row[1] else ''
                email = str(row[2]).strip() if row[2] else ''
                phone = str(row[3]).strip() if row[3] else ''
                register_number = str(row[4]).strip() if row[4] else ''
                department = str(row[5]).strip() if row[5] else ''
                course = str(row[6]).strip() if row[6] else ''
                year = row[7]
                gender = str(row[8]).strip().lower() if row[8] else 'other'

                if not first_name or not email or not register_number or not department or not course or not year:
                    errors.append(f"Row {row_idx}: Missing required fields (First Name, Email, Register Number, Department, Course, Year).")
                    continue
                
                from django.core.validators import validate_email
                from django.core.exceptions import ValidationError
                
                try:
                    validate_email(email)
                except ValidationError:
                    errors.append(f"Row {row_idx}: Invalid email format '{email}'.")
                    continue

                if phone:
                    clean_phone = ''.join(filter(str.isdigit, str(phone)))
                    if len(clean_phone) < 7 or len(clean_phone) > 15:
                        errors.append(f"Row {row_idx}: Phone number '{phone}' is invalid (must be 7-15 digits).")
                        continue
                    phone = clean_phone

                if enforced_dept and department.lower() != enforced_dept.lower():
                    errors.append(f"Row {row_idx}: You can only add students to your own department ({enforced_dept}).")
                    continue

                if Student.objects.filter(email=email).exists():
                    errors.append(f"Row {row_idx}: Email '{email}' already exists.")
                    continue
                    
                if Student.objects.filter(register_number=register_number).exists() or Student.objects.filter(roll_number=register_number).exists():
                    errors.append(f"Row {row_idx}: Register Number '{register_number}' already exists.")
                    continue
                    
                try:
                    year_int = int(year)
                    if year_int < 1 or year_int > 5:
                        errors.append(f"Row {row_idx}: Year must be between 1 and 5.")
                        continue
                except ValueError:
                    errors.append(f"Row {row_idx}: Year must be a number.")
                    continue
                    
                if gender not in ['male', 'female', 'other']:
                    gender = 'other'

                student = Student(
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone=phone,
                    register_number=register_number,
                    roll_number=register_number,
                    department=department,
                    course=course,
                    year=year_int,
                    gender=gender
                )
                student.save()
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: Unexpected error: {str(e)}")

        return Response({
            'success': True,
            'created_count': created_count,
            'errors': errors
        })
