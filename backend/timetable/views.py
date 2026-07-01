from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from .models import TimetableEntry
from .serializers import TimetableEntrySerializer

class TimetableViewSet(viewsets.ModelViewSet):
    queryset = TimetableEntry.objects.all()
    serializer_class = TimetableEntrySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['course', 'subject', 'faculty_name', 'day']

    def get_queryset(self):
        qs = super().get_queryset()
        p = self.request.query_params
        if p.get('course'):        qs = qs.filter(course=p['course'])
        if p.get('department'):    qs = qs.filter(department=p['department'])
        if p.get('semester'):      qs = qs.filter(semester=p['semester'])
        if p.get('day'):           qs = qs.filter(day=p['day'])
        if p.get('academic_year'): qs = qs.filter(academic_year=p['academic_year'])
        return qs

import openpyxl
import datetime
from rest_framework.views import APIView
from rest_framework.response import Response

class TimetableBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)
            
        if not file.name.endswith('.xlsx'):
            return Response({'error': 'Please upload a valid .xlsx file'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=400)

        created_count = 0
        errors = []
        
        valid_days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx <= 2:
                continue
                
            if not any(row):
                continue
                
            try:
                # 0: Course, 1: Dept, 2: Sem, 3: Day, 4: Period, 5: Start Time, 6: End Time, 
                # 7: Subject, 8: Sub Code, 9: Faculty, 10: Room, 11: Academic Year
                course = str(row[0]).strip() if row[0] else ''
                department = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                semester_raw = row[2] if len(row) > 2 else None
                day = str(row[3]).strip().lower() if len(row) > 3 and row[3] else ''
                period_raw = row[4] if len(row) > 4 else None
                start_time_raw = row[5] if len(row) > 5 else None
                end_time_raw = row[6] if len(row) > 6 else None
                subject = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                subject_code = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                faculty_name = str(row[9]).strip() if len(row) > 9 and row[9] else ''
                room = str(row[10]).strip() if len(row) > 10 and row[10] else ''
                academic_year = str(row[11]).strip() if len(row) > 11 and row[11] else ''
                
                if not course or not day or period_raw is None or not start_time_raw or not end_time_raw or not subject:
                    errors.append(f"Row {row_idx}: Missing required fields (Course, Day, Period, Start Time, End Time, Subject).")
                    continue
                    
                if day not in valid_days:
                    errors.append(f"Row {row_idx}: Invalid day '{day}'. Allowed: Monday-Saturday.")
                    continue
                    
                try:
                    semester = int(semester_raw) if semester_raw else 1
                    period = int(period_raw)
                except ValueError:
                    errors.append(f"Row {row_idx}: Semester and Period must be valid numbers.")
                    continue
                    
                # Time parsing
                def parse_time(val):
                    if isinstance(val, datetime.time): return val
                    if isinstance(val, datetime.datetime): return val.time()
                    if isinstance(val, str):
                        try:
                            # Try simple formats
                            val = val.strip()
                            if len(val) == 4 and val.isdigit(): # e.g. 0900
                                return datetime.time(int(val[:2]), int(val[2:]))
                            parts = val.replace(':', ' ').replace('.', ' ').split()
                            h = int(parts[0])
                            m = int(parts[1]) if len(parts) > 1 else 0
                            s = int(parts[2]) if len(parts) > 2 else 0
                            return datetime.time(h, m, s)
                        except: pass
                    return None
                    
                start_time = parse_time(start_time_raw)
                end_time = parse_time(end_time_raw)
                
                if not start_time or not end_time:
                    errors.append(f"Row {row_idx}: Invalid time format. Please use HH:MM format.")
                    continue
                    
                # Check unique constraint
                if TimetableEntry.objects.filter(course=course, semester=semester, day=day, period=period, academic_year=academic_year).exists():
                    errors.append(f"Row {row_idx}: Entry for {course} Sem {semester} on {day.title()} Period {period} already exists.")
                    continue
                    
                TimetableEntry.objects.create(
                    course=course, department=department, semester=semester, day=day, period=period,
                    start_time=start_time, end_time=end_time, subject=subject, subject_code=subject_code,
                    faculty_name=faculty_name, room=room, academic_year=academic_year
                )
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: Unexpected error: {str(e)}")

        return Response({
            'success': True,
            'created_count': created_count,
            'errors': errors
        })

from django.http import HttpResponse
from rest_framework.permissions import AllowAny

class TimetableBulkUploadSampleView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"

        headers = ['Course', 'Department', 'Semester', 'Day', 'Period', 'Start Time', 'End Time', 'Subject', 'Subject Code', 'Faculty Name', 'Room', 'Academic Year']
        ws.append(headers)

        example = ['B.Tech', 'Computer Science', 1, 'monday', 1, '09:00', '09:50', 'Data Structures', 'CS101', 'John Doe', 'Room 101', '2024-25']
        ws.append(example)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=timetable_bulk_upload_sample.xlsx'
        wb.save(response)
        return response

from django.http import HttpResponse
from rest_framework.permissions import AllowAny

class TimetableBulkUploadSampleView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"

        headers = ['Course', 'Department', 'Semester', 'Day', 'Period', 'Start Time', 'End Time', 'Subject', 'Subject Code', 'Faculty Name', 'Room', 'Academic Year']
        ws.append(headers)

        example = ['B.Tech', 'Computer Science', 1, 'monday', 1, '09:00', '09:50', 'Data Structures', 'CS101', 'John Doe', 'Room 101', '2024-25']
        ws.append(example)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=timetable_bulk_upload_sample.xlsx'
        wb.save(response)
        return response
