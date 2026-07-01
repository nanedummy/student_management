from rest_framework import generics, filters
from config.pagination import StandardResultsSetPagination
from .models import Faculty
from .serializers import FacultySerializer
from accounts.permissions import IsAdminOrFacultyReadOnly, IsAdmin


class FacultyListCreateView(generics.ListCreateAPIView):
    serializer_class = FacultySerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAdminOrFacultyReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'email', 'department', 'designation']

    def get_queryset(self):
        queryset = Faculty.objects.all()
        
        course = self.request.query_params.get('course')
        if course:
            queryset = queryset.filter(course=course)
            
        department = self.request.query_params.get('department')
        if department:
            queryset = queryset.filter(department=department)
            
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        designation = self.request.query_params.get('designation')
        if designation:
            queryset = queryset.filter(designation=designation)
            
        return queryset.order_by('-created_at')


class FacultyDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Faculty.objects.all()
    serializer_class = FacultySerializer
    permission_classes = [IsAdminOrFacultyReadOnly]

import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response

class FacultyBulkUploadView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)
            
        if not file.name.endswith('.xlsx'):
            return Response({'error': 'Please upload a valid .xlsx file'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=400)

        created_count = 0
        errors = []

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx <= 2: # Skip headers and example
                continue
                
            if not any(row):
                continue
                
            try:
                # 0: First Name, 1: Last Name, 2: Email, 3: Phone, 4: Dept, 5: Course, 6: Designation, 7: Qual, 8: Exp
                first_name = str(row[0]).strip() if row[0] else ''
                last_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                email = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                phone = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                department = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                course = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                designation = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                qualification = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                experience_raw = row[8] if len(row) > 8 else 0
                
                if not first_name or not email or not department or not designation:
                    errors.append(f"Row {row_idx}: Missing required fields (First Name, Email, Department, Designation).")
                    continue
                    
                if Faculty.objects.filter(email=email).exists():
                    errors.append(f"Row {row_idx}: Email '{email}' already exists.")
                    continue
                    
                experience = 0
                if experience_raw:
                    try: experience = int(experience_raw)
                    except: pass
                    
                Faculty.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone=phone,
                    department=department,
                    course=course,
                    designation=designation,
                    qualification=qualification,
                    experience=experience,
                    status='active'
                )
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: Unexpected error: {str(e)}")

        return Response({
            'success': True,
            'created_count': created_count,
            'errors': errors
        })

from django.http import HttpResponse
from rest_framework.permissions import AllowAny

class FacultyBulkUploadSampleView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Faculty"

        headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Department', 'Course', 'Designation', 'Qualification', 'Experience (Years)']
        ws.append(headers)

        example = ['Dr. John', 'Doe', 'johndoe@example.com', '9876543210', 'Computer Science', 'B.Tech', 'Professor', 'Ph.D', 10]
        ws.append(example)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=faculty_bulk_upload_sample.xlsx'
        wb.save(response)
        return response
